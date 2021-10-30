"""
Y.Wang
Date: 04.10.2021
"""
import pandas as pd
import openpyxl
import xlwt
import xlrd
import datetime
from xlwt import easyxf

Bygma_data = pd.read_csv(r'C:\Users\User\Desktop\temp\record.csv')
Husa_data = pd.read_csv(r'C:\Users\User\Desktop\temp\record (1).csv')
Samkaup_data = pd.read_csv(r'C:\Users\User\Desktop\temp\record (2).csv')
Bunnpris_data = pd.read_excel(r'C:\Users\User\Desktop\temp\updateRecord.xlsx')
Bunnpris_data.to_csv(r'C:\Users\User\Desktop\temp\updateRecord.csv')
Magasin_data = pd.read_csv(r'C:\Users\User\Desktop\temp\record (3).csv')

x = Bygma_data['Update Result']
y = Husa_data['Status']
z = Samkaup_data['Status']
w = Bunnpris_data['Result']
v = Magasin_data['Status']
print("###################Bygma######################")
#A=[x,y,z,w]
#for i in enumerate (A):
data1=[]
data2=[]
data3=[]
data4=[]
data5=[]
data6=[]
data7=[]
for index,item in enumerate (x):
    if item=="Update Succeeded" :
        data1.append(item)
    elif item=="Cancel" :
        data2.append(item)
    elif item=="ESL Failure" :
        data3.append(item)
    elif item=="ESL Offline" :
        data4.append(item)
    elif item=="Data Error" :
        data5.append(item)
    elif item=="Pending Processing" :
        data6.append(item)
    elif item=="Template error" :
        data7.append(item)
print(len(x))
print(len(data1))
print(len(data2))
print(len(data3))
print(len(data4))
print(len(data5))
print(len(data6))
print(len(data7))
if len(data1)+len(data2)+len(data3)+len(data4)+len(data5)+len(data6)+len(data7)!=len(x):
    print("new Update Result type")
print("#################Husa#########################")
data1=[]
data2=[]
data3=[]
data4=[]
data5=[]
data6=[]
data7=[]
for index,item in enumerate (y):
    if item=="Succeeded" :
        data1.append(item)
    elif item=="AP Offline" :
        data2.append(item)
    elif item=="Cancel" :
        data3.append(item)
    elif item=="Offline" :
        data4.append(item)
    elif item=="Offline w/o HB" :
        data5.append(item)
    elif item=="Template Error" :
        data6.append(item)
    elif item=="Timeout" :
        data7.append(item)
print(len(y))
print(len(data1))
print(len(data2))
print(len(data3))
print(len(data4))
print(len(data5))
print(len(data6))
print(len(data7))
if len(data1)+len(data2)+len(data3)+len(data4)+len(data5)+len(data6)+len(data7)!=len(y):
    print("new Status type")
print("##############Samkaup#########################")
data1=[]
data2=[]
data3=[]
data4=[]
data5=[]
data6=[]
data7=[]
data8=[]
for index,item in enumerate (z):
    if item=="Succeeded" :
        data1.append(item)
    elif item=="AP Offline" :
        data2.append(item)
    elif item=="Cancel" :
        data3.append(item)
    elif item=="ESL Error" :
        data4.append(item)
    elif item=="Offline" :
        data5.append(item)
    elif item=="Offline w/o HB" :
        data6.append(item)
    elif item=="Template error" :
        data7.append(item)
    elif item=="Timeout" :
        data8.append(item)
print(len(z))
print(len(data1))
print(len(data2))
print(len(data3))
print(len(data4))
print(len(data5))
print(len(data6))
print(len(data7))
print(len(data8))
if len(data1)+len(data2)+len(data3)+len(data4)+len(data5)+len(data6)+len(data7)+len(data8)!=len(z):
    print("new Status type")
print("##############Bunnpris#########################")
Sam_data1=[]
Sam_data2=[]
Sam_data3=[]
Sam_data4=[]
Sam_data5=[]
Sam_data6=[]
Sam_data7=[]
Sam_data8=[]
for index,item in enumerate (w):
    if item=="Update Success" :
        Sam_data1.append(item)
    elif item=="Blacklist" :
        Sam_data2.append(item)
    elif item=="Cancel" :
        Sam_data3.append(item)
    elif item=="ESL Error" :
        Sam_data4.append(item)
    elif item=="ESL Offline" :
        Sam_data5.append(item)
    elif item=="Station Offline" :
        Sam_data6.append(item)
    elif item=="Template Error" :
        Sam_data7.append(item)
    elif item=="Update Failed" :
        Sam_data8.append(item)
print(len(w))
print(len(Sam_data1))
print(len(Sam_data2))
print(len(Sam_data3))
print(len(Sam_data4))
print(len(Sam_data5))
print(len(Sam_data6))
print(len(Sam_data7))
print(len(Sam_data8))
if len(Sam_data1)+len(Sam_data2)+len(Sam_data3)+len(Sam_data4)+len(Sam_data5)+len(Sam_data6)+len(Sam_data7)+len(Sam_data8)!=len(w):
    print("new Result type")
print("##############Magasin#########################")
data1=[]
data2=[]
data3=[]
data4=[]
data5=[]
data6=[]
data7=[]
data8=[]
for index,item in enumerate (z):
    if item=="Succeeded" :
        data1.append(item)
    elif item=="AP Offline" :
        data2.append(item)
    elif item=="Cancel" :
        data3.append(item)
    elif item=="ESL Error" :
        data4.append(item)
    elif item=="Offline" :
        data5.append(item)
    elif item=="Offline w/o HB" :
        data6.append(item)
    elif item=="Template error" :
        data7.append(item)
    elif item=="Timeout" :
        data8.append(item)
print(len(z))
print(len(data1))
print(len(data2))
print(len(data3))
print(len(data4))
print(len(data5))
print(len(data6))
print(len(data7))
print(len(data8))
if len(data1)+len(data2)+len(data3)+len(data4)+len(data5)+len(data6)+len(data7)+len(data8)!=len(z):
    print("new Status type")

workbook = xlwt.Workbook() #create a excel file
worksheet = workbook.add_sheet('Tabelle1') #create a table
'''
style    light grey background, middle, bold
style1   original
style2   time
style3   light yellow background, middle, bold, #title
style4   light blue, middle
style5   light green, middle
'''

style = xlwt.XFStyle() # Create Style
alignment = xlwt.Alignment() # Create Alignment
alignment.horz = xlwt.Alignment.HORZ_CENTER # May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
alignment.vert = xlwt.Alignment.VERT_CENTER # May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
style.alignment = alignment # Add Alignment to Style

pattern = xlwt.Pattern() # Create the Pattern
pattern.pattern = xlwt.Pattern.SOLID_PATTERN # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
pattern.pattern_fore_colour = 22 # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
style.pattern = pattern # Add Pattern to Style

font = xlwt.Font() # 为样式创建字体
font.name = 'Times New Roman'
font.bold = True # 黑体
#font.underline = True # 下划线
#font.italic = True # 斜体字
style.font = font # 设定样式
'''
border=xlwt.Borders() #实例化边框
border.top=xlwt.Borders.THIN #设置上边框为实线
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
#border.left=xlwt.Borders.DOTTED #设置左边框为点状线
#border.right=xlwt.Borders.NO_LINE #设置右边框为默认值
style.borders=border
#xw.Borders？？ 查看说明文档
'''
style1 = xlwt.XFStyle()
border=xlwt.Borders()
border.top=xlwt.Borders.THIN
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
style1.borders=border

style2 = xlwt.XFStyle()
style2.num_format_str = 'D-MMM-YY' # Other options: D-MMM-YY, D-MMM, MMM-YY, h:mm, h:mm:ss, h:mm, h:mm:ss, M/D/YY h:mm, mm:ss, [h]:mm:ss, mm:ss.0

border=xlwt.Borders()
border.top=xlwt.Borders.THIN
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
style2.borders=border

style3 = xlwt.XFStyle() # Create the Pattern

pattern = xlwt.Pattern() # Create the Pattern
pattern.pattern = xlwt.Pattern.SOLID_PATTERN # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
pattern.pattern_fore_colour = 26 # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
style3.pattern = pattern # Add Pattern to Style

alignment = xlwt.Alignment() # Create Alignment
alignment.horz = xlwt.Alignment.HORZ_CENTER # May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
alignment.vert = xlwt.Alignment.VERT_CENTER # May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
style3.alignment = alignment # Add Alignment to Style

font = xlwt.Font() # 为样式创建字体
font.name = 'Times New Roman'
font.bold = True
font.height = 20 * 18  # 字体大小  20位衡量单位
style3.font = font # 设定样式
'''
border=xlwt.Borders()
border.top=xlwt.Borders.THIN
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
style3.borders=border
'''
style4 = xlwt.XFStyle() # Create the Pattern

pattern = xlwt.Pattern() # Create the Pattern
pattern.pattern = xlwt.Pattern.SOLID_PATTERN # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
pattern.pattern_fore_colour = 44 # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
style4.pattern = pattern # Add Pattern to Style

alignment = xlwt.Alignment() # Create Alignment
alignment.horz = xlwt.Alignment.HORZ_CENTER # May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
alignment.vert = xlwt.Alignment.VERT_CENTER # May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
style4.alignment = alignment # Add Alignment to Style

border=xlwt.Borders()
border.top=xlwt.Borders.THIN
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
style4.borders=border

style5 = xlwt.XFStyle() # Create the Pattern

pattern = xlwt.Pattern() # Create the Pattern
pattern.pattern = xlwt.Pattern.SOLID_PATTERN # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
pattern.pattern_fore_colour = 42 # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
style5.pattern = pattern # Add Pattern to Style

alignment = xlwt.Alignment() # Create Alignment
alignment.horz = xlwt.Alignment.HORZ_CENTER # May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
alignment.vert = xlwt.Alignment.VERT_CENTER # May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
style5.alignment = alignment # Add Alignment to Style

font = xlwt.Font() # 为样式创建字体
font.name = 'Times New Roman'
font.bold = True
font.height = 20 * 16  # 字体大小  20位衡量单位
style5.font = font # 设定样式
'''
border=xlwt.Borders()
border.top=xlwt.Borders.THIN
border.left=xlwt.Borders.THIN
border.right=xlwt.Borders.THIN
border.bottom=xlwt.Borders.THIN
style5.borders=border
'''
worksheet.write(0, 0, "Daily Store Inspection Report", style3)
worksheet.merge(0, 0, 0, 3) # Merges row 0's columns 0 through 3.
worksheet.write(7, 2, "Good", style5)
worksheet.merge(7, 22, 2, 2)
worksheet.merge(7, 22, 3, 3)

worksheet.merge(2, 2, 1, 3,style=style)
worksheet.merge(3, 3, 1, 3)
worksheet.merge(4, 4, 1, 3)
#merged = worksheet.merged_cells
#worksheet.merged_cells('A1:D1')
worksheet.row(0).height_mismatch = True
worksheet.row(0).height = 20 * 40
worksheet.col(0).width = 6666
worksheet.col(2).width = 4000
worksheet.col(3).width = 6666
worksheet.write(1, 0, "Store Information", style)
worksheet.merge(1, 1, 0, 3)
worksheet.write(2, 0, "Store", style1)
worksheet.write(2, 1, "Samkaup", style2)
worksheet.write(3, 0, "Inspected by", style1)
worksheet.write(3, 1, "Y.Wang", style2)
worksheet.write(4, 0, "Inspection time", style1)
worksheet.write(4, 1, datetime.datetime.now(), style2)
worksheet.write(5, 0, "System Report", style)
worksheet.merge(5, 5, 0, 3)
worksheet.write(6, 0, "Type", style4)
worksheet.write(6, 1, "Amount", style4)
worksheet.write(6, 2, "System Status", style4)
worksheet.write(6, 3, "Detail Info", style4)
worksheet.write(7, 0, "Total ESL Amount", style1)
worksheet.write(7, 1, "7996", style1)
worksheet.write(8, 0, "Total ELS - 2.13 Freezer", style1)
worksheet.write(8, 1, "159", style1)
worksheet.write(9, 0, "Total ELS - 2.13 Freezer", style1)
worksheet.write(9, 1, "159", style1)
worksheet.write(10, 0, "Total ELS - 2.13 Normal", style1)
worksheet.write(10, 1, "6879", style1)
worksheet.write(11, 0, "Total ESL - 1.54 Normal", style1)
worksheet.write(11, 1, "800", style1)
worksheet.write(12, 0, "Total ESL - 4.20 Normal", style1)
worksheet.write(12, 1, "58", style1)
worksheet.write(13, 0, "Total update ESL", style1)
worksheet.write(13, 1, len(w), style1)
worksheet.write(14, 0, "Succeeded", style1)
worksheet.write(14, 1, len(Sam_data1), style1)
worksheet.write(15, 0, "AP Offline", style1)
worksheet.write(15, 1, len(Sam_data2), style1)
worksheet.write(16, 0, "Cancel", style1)
worksheet.write(16, 1, len(Sam_data3), style1)
worksheet.write(17, 0, "ESL Error", style1)
worksheet.write(17, 1, len(Sam_data4), style1)
worksheet.write(18, 0, "Offline", style1)
worksheet.write(18, 1, len(Sam_data5), style1)
worksheet.write(19, 0, "Offline w/o HB", style1)
worksheet.write(19, 1, len(Sam_data6), style1)
worksheet.write(20, 0, "Template error", style1)
worksheet.write(20, 1, len(Sam_data7), style1)
worksheet.write(21, 0, "Timeout", style1)
worksheet.write(21, 1, len(Sam_data8), style1)
worksheet.write(22, 0, "Other Error", style1)
worksheet.write(22, 1, len(w)-(len(Sam_data1)+len(Sam_data2)+len(Sam_data3)+len(Sam_data4)+len(Sam_data5)+len(Sam_data6)+len(Sam_data7)+len(Sam_data8)), style1)

worksheet.write(22, 0, "Other Error", style1)
worksheet.write(22, 0, "Other Error", style1)
worksheet.write(22, 0, "Other Error", style1)
worksheet.write(22, 0, "Other Error", style1)


workbook.save('Samkaup_Daily Store Inspection Report.xls')


#wb=openpyxl.load_workbook(r'C:\Users\User\Desktop\temp\Samkaup_daily Store Inspection Report_HANSHOW.xlsx')
#ws=wb.worksheets[0]
#worksheet.cell(row=12, column=3).value = '123'
#ws.write(12,5,"123", style)
#print(openpyxl.__version__)